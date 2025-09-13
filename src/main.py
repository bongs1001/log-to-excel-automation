# ===== 1) 임포트 & 경로 =====
import os, re, json, time, base64
from typing import Dict, Any, List
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
import google.generativeai as genai
from google.generativeai import GenerationConfig

# 레포 구조 기준 (src/ 상위가 프로젝트 루트)
BASE       = os.path.dirname(os.path.dirname(__file__))                           
INPUT_DIR  = os.path.join(BASE, "data")
OUT_DIR    = os.path.join(BASE, "outputs")
EXCEL_PATH = os.path.join(OUT_DIR, "log_to_excel_output.xlsx")
SHEET_NAME = "Sheet1"

os.makedirs(OUT_DIR, exist_ok=True)
assert os.path.isdir(INPUT_DIR), f"입력 폴더 없음: {INPUT_DIR}"

# ===== 2) 모델 설정 =====
load_dotenv()
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
SYSTEM = "너는 회의록을 구조적 JSON으로 추출하는 비서이다. 반드시 유효한 JSON만 출력한다."
gmodel_json = genai.GenerativeModel(
    "gemini-2.5-flash", 
    system_instruction=SYSTEM,
    generation_config=GenerationConfig(
        temperature=0.2,
        max_output_tokens=2048,
        response_mime_type="application/json" 
    )
)

gmodel_text = genai.GenerativeModel(         # 백업: 텍스트 모드
    "gemini-1.5-flash",
    system_instruction=SYSTEM,
    generation_config=GenerationConfig(
        temperature=0.2,
        max_output_tokens=2048
    )
)

# ===== 3) 응답 파싱 유틸 =====
def extract_response_str(resp):
    if getattr(resp, "text", None):
        return resp.text
    for c in getattr(resp, "candidates", []) or []:
        content = getattr(c, "content", None)
        parts = getattr(content, "parts", None) if content else None
        if not parts: 
            continue
        for p in parts:
            if hasattr(p, "text") and p.text:
                return p.text
            if hasattr(p, "inline_data") and getattr(p.inline_data, "data", None):
                try:
                    return base64.b64decode(p.inline_data.data).decode("utf-8","ignore")
                except Exception:
                    pass
    raise RuntimeError("모델 응답에서 텍스트/JSON part 없음")

def gemini_call_dual(prompt_user: str, retry_each: int = 1) -> str:
    # (a) JSON 강제 모델
    for attempt in range(retry_each+1):
        try:
            return extract_response_str(gmodel_json.generate_content(prompt_user))
        except Exception as e:
            last_json_err = e
            time.sleep(1.0 * (attempt + 1))
    # (b) 텍스트 모드
    for attempt in range(retry_each+1):
        try:
            return extract_response_str(gmodel_text.generate_content(prompt_user))
        except Exception as e:
            last_text_err = e
            time.sleep(1.0 * (attempt + 1))
    # 둘 다 실패
    raise RuntimeError(f"GENERATION_FAILED: json={last_json_err} / text={last_text_err}")

# ===== 4) JSON/마커 파서 =====
def parse_json_loose(text: str):
    t = text.strip().lstrip("\ufeff")
    t = re.sub(r"^```json", "", t, flags=re.I).strip()
    t = re.sub(r"^```", "", t).strip()
    t = re.sub(r"```$", "", t).strip()
    m = re.search(r"(\{.*\}|\[.*\])\s*$", t, flags=re.S)
    if not m:
        raise ValueError("JSON 본문 미검출")
    return json.loads(m.group(1))

def parse_marked_blocks(text: str):
    def grab(tag):
        m = re.search(rf"<<{tag}>>\s*(.*?)\s*(?=<<|$)", text, flags=re.S)
        return (m.group(1).strip() if m else "")
    return {
        "date": grab("DATE"),
        "mentor": grab("MENTOR"),
        "content": grab("CONTENT"),
        "mentoring": grab("MENTORING"),
    }

# ===== 5) 보조 유틸 =====
def sanitize_transcript(t: str, max_chars: int = 115000) -> str:
    t = t[:max_chars]
    t = re.sub(r"\b\d{2,4}-\d{3,4}-\d{4}\b", "[PHONE]", t)
    t = re.sub(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", "[EMAIL]", t)
    return t

SCHEMA = """스키마:{
 "session_datetime":"YYYY-MM-DD HH:MM"|"unknown",
 "mentor":"string","attendees":["string",...],
 "summary_title":"핵심주제 한줄","highlights":["불릿1","불릿2","불릿3","불릿4","불릿5"],
 "decisions":["결정1",...],
 "action_items":[{"assignee":"이름","task":"할일","due":"YYYY-MM-DD|null"}],
 "risks":["리스크1",...],"next_plan":"다음 계획/요청","tags":["#태그1","#태그2"]
}
규칙: JSON 외 텍스트/마크다운/코드블록 금지
"""

# 6) 요약: JSON → 실패 시 텍스트 → 실패 시 마커(구제)
def summarize_block(txt: str) -> Dict[str, Any]:
    t = sanitize_transcript(txt)
    user = f"아래 녹취록을 요약해. {SCHEMA}\n녹취록:\n{t}"
    try:
        out = gemini_call_dual(user, retry_each=1)
        return parse_json_loose(out)
    except Exception:
        prompt = f"""아래 녹취록을 요약하되, 다음 형식만 출력하고 그 외 문구는 금지.

<<DATE>>
YYYY-MM-DD HH:MM 또는 unknown

<<MENTOR>>
이름(모르면 공란)

<<CONTENT>>
(핵심주제 한 줄)
- 하이라이트1
- 하이라이트2
- 하이라이트3

<<MENTORING>>
■ 결정사항
- ...
■ 액션아이템
- [담당자] 할일 (기한: YYYY-MM-DD|null)
■ 리스크/이슈
- ...
■ 다음 계획
...

녹취록:
{t}
"""
        out2 = gemini_call_dual(prompt, retry_each=1)
        blocks = parse_marked_blocks(out2)
        lines = [ln.strip() for ln in (blocks["content"] or "").splitlines() if ln.strip()]
        title = lines[0] if lines else ""
        highs = [ln.lstrip("- ").strip() for ln in lines[1:] if ln.startswith("-")]
        return {
            "session_datetime": (blocks["date"] or "unknown"),
            "mentor": blocks["mentor"],
            "attendees": [],
            "summary_title": title,
            "highlights": highs,
            "decisions": [],
            "action_items": [],
            "risks": [],
            "next_plan": "",
            "tags": []
        }

# ===== 7) 엑셀 유틸 =====
def find_header_cols(ws) -> Dict[str, int]:
    header_map = {}
    for col in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str):
            header_map[v.strip()] = col
    return header_map

def next_empty_row(ws, header_row:int=1) -> int:
    r = header_row + 1
    while True:
        values = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if all(v in (None, "") for v in values):
            return r
        r += 1

def save_wb_with_retry(wb, path, tries=5, wait=2.0):
    for i in range(tries):
        try:
            wb.save(path)
            return
        except PermissionError:
            if i == 0:
                print(f"[WARN] 엑셀 파일이 열려있습니다. 닫아주세요: {path}")
            time.sleep(wait)
    raise PermissionError(f"저장 실패(파일 잠금 지속): {path}")

def build_cells_for_template(struct:Dict[str,Any]) -> Dict[str,str]:
    ses = struct.get("session_datetime") or "unknown"
    mentor = struct.get("mentor") or ""
    date_block = ses if ses != "unknown" else ""
    if mentor:
        date_block = f"{date_block}\n{mentor}".strip()
    highs = struct.get("highlights") or []
    content = (struct.get("summary_title") or "").strip()
    if highs:
        content = (content + "\n" + "\n".join(highs)).strip()

    decisions = struct.get("decisions") or []
    actions   = struct.get("action_items") or []
    risks     = struct.get("risks") or []
    nextp     = struct.get("next_plan") or ""

    def fmt_actions(items):
        lines = []
        for it in items:
            a = it.get("assignee") or "-"
            t = it.get("task") or "-"
            d = it.get("due") or "-"
            lines.append(f"- [{a}] {t} (기한: {d})")
        return "\n".join(lines)

    mentoring_par = ""
    if decisions: mentoring_par += "■ 결정사항\n" + "\n".join(f"- {d}" for d in decisions) + "\n\n"
    if actions:   mentoring_par += "■ 액션아이템\n" + fmt_actions(actions) + "\n\n"
    if risks:     mentoring_par += "■ 리스크/이슈\n" + "\n".join(f"- {r}" for r in risks) + "\n\n"
    if nextp:     mentoring_par += "■ 다음 계획\n" + nextp

    return {
        "회차": "",
        "일자 및 시간": date_block,
        "장 소": "팀즈 화상",
        "구분": "주제 멘토링",
        "내용": content,
        "멘토링": mentoring_par.strip()
    }

def write_to_template(excel_path:str, sheet_name:str, struct:Dict[str,Any], filename:str) -> int:
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=["원본파일명","회차","일자 및 시간","장 소","구분","내용","멘토링"])
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    header_map = find_header_cols(ws)
    row = next_empty_row(ws, header_row=1)
    cells = build_cells_for_template(struct)
    cells["원본파일명"] = filename  # ✅ 맨 앞 열

    # 회차 자동 증가
    if "회차" in header_map:
        c = header_map["회차"]
        prev = ws.cell(row=row-1, column=c).value
        try:
            seq = int(prev) + 1 if isinstance(prev, (int, float)) else 1
        except Exception:
            seq = 1
        ws.cell(row=row, column=c, value=seq)

    # 값 쓰기
    for k, v in cells.items():
        if k in header_map:
            ws.cell(row=row, column=header_map[k], value=v)

    save_wb_with_retry(wb, excel_path)
    return row

# ===== 8) 실행 =====
processed, errors = 0, []
for f in os.listdir(INPUT_DIR):
    if not f.lower().endswith(".txt"):
        continue
    p = os.path.join(INPUT_DIR, f)
    try:
        with open(p, "r", encoding="utf-8") as fh:
            transcript = fh.read()
        struct = summarize_block(transcript)
        r = write_to_template(EXCEL_PATH, SHEET_NAME, struct, filename=f)
        processed += 1
        print(f"[OK] {f} -> row {r}")
    except Exception as e:
        errors.append((f, str(e)))
        print(f"[ERR] {f}: {e}")

print(f"\n완료: {processed}건 처리")
if errors:
    print("오류 목록:")
    for f, msg in errors:
        print(" -", f, ":", msg)
